from __future__ import annotations

import argparse
import hashlib
import json
import logging
import random
import re
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook, load_workbook

from .config import load_config, project_path

USER_AGENT = "ow-elo-backfill-bot/1.0 (private research tool; not for redistribution)"
PORTAL_PATH = "/overwatch/Portal:Tournaments"
BACKFILL_END_YEAR = 2026
MATCH_PREFIX_HEADERS = ["match_date", "event_name", "stage"]
MATCH_SUFFIX_HEADERS = [
    "score_a",
    "score_b",
    "best_of",
    "winning_side",
    "notes",
    "status",
    "import_error",
    "row_hash",
    "roster_confidence",
]
MATCH_HEADERS = [*MATCH_PREFIX_HEADERS, *[f"player_a{i}" for i in range(1, 12)], *[f"player_b{i}" for i in range(1, 12)], *MATCH_SUFFIX_HEADERS]
ALIAS_HEADERS = ["canonical_player", "alias_text", "source", "approved"]
IMPORT_LOG_HEADERS = ["import_time", "rows_seen", "rows_imported", "rows_rejected", "notes"]
HTML_PARSER = "lxml"


@dataclass
class Tournament:
    title: str
    path: str
    tier: str = ""
    date_hint: str = ""
    date_hint_source: str = ""


@dataclass
class ParsedMatch:
    match_date: str
    event_name: str
    stage: str
    team_a: str
    team_b: str
    score_a: int
    score_b: int
    best_of: int
    source_url: str
    players_a: list[str] = field(default_factory=list)
    players_b: list[str] = field(default_factory=list)
    roster_confidence: str = "missing"
    notes: list[str] = field(default_factory=list)


@dataclass
class Summary:
    tournaments_processed: int = 0
    matches_found: int = 0
    exact: int = 0
    inferred: int = 0
    partial: int = 0
    missing: int = 0
    rows_ready: int = 0
    rows_draft: int = 0
    rows_skipped_dup: int = 0
    rows_skipped_scope: int = 0
    cache_hits: int = 0
    unresolved_players: int = 0
    parse_warnings: int = 0
    errors: int = 0


class LiquipediaBot:
    def __init__(
        self,
        mode: str,
        pages: list[str] | None = None,
        html_files: list[str] | None = None,
        terminal_doc: str = "",
    ) -> None:
        self.mode = mode
        self.page_paths = pages or []
        self.html_files = [Path(path) for path in (html_files or [])]
        self.terminal_doc = project_path(terminal_doc) if terminal_doc else None
        self.config = load_config()
        self.base_url = self.config["liquipedia"]["base_url"].rstrip("/")
        self.rate_limit = max(2.0, float(self.config["liquipedia"]["rate_limit_seconds"]))
        self.jitter_min = float(self.config["liquipedia"].get("jitter_min_seconds", -1.0))
        self.jitter_max = float(self.config["liquipedia"].get("jitter_max_seconds", 15.0))
        self.jitter_mode = float(self.config["liquipedia"].get("jitter_mode_seconds", 8.0))
        self.max_retries = int(self.config["liquipedia"]["max_retries"])
        self.cache_dir = project_path(self.config["liquipedia"]["cache_dir"])
        self.cache_ttl = timedelta(days=int(self.config["liquipedia"]["cache_ttl_days"]))
        self.state_file = project_path(self.config["liquipedia"]["state_file"])
        self.workbook_path = project_path(self.config["paths"]["workbook"])
        self.log_path = project_path(self.config["paths"]["logs"]) / "liquipedia_bot.log"
        self.last_request_at = 0.0
        self.summary = Summary()
        self.state = self.load_state()
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})
        self.setup_logging()

    def setup_logging(self) -> None:
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        handlers: list[logging.Handler] = [logging.FileHandler(self.log_path, encoding="utf-8"), logging.StreamHandler()]
        if self.terminal_doc:
            self.terminal_doc.parent.mkdir(parents=True, exist_ok=True)
            handlers.append(logging.FileHandler(self.terminal_doc, mode="w", encoding="utf-8"))
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s %(levelname)s %(message)s",
            handlers=handlers,
            force=True,
        )
        if self.terminal_doc:
            logging.info("Terminal transcript will be written to %s", self.terminal_doc)

    def load_state(self) -> dict[str, Any]:
        if self.state_file.exists():
            try:
                return json.loads(self.state_file.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                pass
        return {
            "last_run_timestamp": None,
            "processed_tournament_urls": [],
            "total_matches_written": 0,
            "total_matches_skipped": 0,
        }

    def save_state(self, rows_written: int, rows_skipped: int, processed: list[str]) -> None:
        if self.mode == "dry-run":
            return
        self.state_file.parent.mkdir(parents=True, exist_ok=True)
        processed_urls = sorted(set(self.state.get("processed_tournament_urls", []) + processed))
        self.state.update(
            {
                "last_run_timestamp": datetime.now(timezone.utc).isoformat(),
                "processed_tournament_urls": processed_urls,
                "total_matches_written": int(self.state.get("total_matches_written", 0)) + rows_written,
                "total_matches_skipped": int(self.state.get("total_matches_skipped", 0)) + rows_skipped,
            }
        )
        self.state_file.write_text(json.dumps(self.state, indent=2), encoding="utf-8")

    def cache_path(self, url: str) -> Path:
        digest = hashlib.sha256(url.encode("utf-8")).hexdigest()
        return self.cache_dir / f"{digest}.html"

    def fetch(self, path_or_url: str) -> str | None:
        url = urljoin(self.base_url, path_or_url)
        cache_path = self.cache_path(url)
        if cache_path.exists() and datetime.now() - datetime.fromtimestamp(cache_path.stat().st_mtime) < self.cache_ttl:
            self.summary.cache_hits += 1
            return cache_path.read_text(encoding="utf-8", errors="replace")

        self.cache_dir.mkdir(parents=True, exist_ok=True)
        elapsed = time.monotonic() - self.last_request_at
        request_delay = self.request_delay_seconds()
        if elapsed < request_delay:
            time.sleep(request_delay - elapsed)

        for attempt in range(self.max_retries + 1):
            try:
                response = self.session.get(url, timeout=30)
                self.last_request_at = time.monotonic()
            except requests.RequestException as exc:
                logging.warning("Connection error for %s: %s", url, exc)
                if attempt >= self.max_retries:
                    self.summary.errors += 1
                    return None
                time.sleep(10)
                continue

            if response.status_code == 200:
                cache_path.write_text(response.text, encoding="utf-8")
                return response.text
            if response.status_code == 404:
                logging.info("Page not found: %s", url)
                return None
            if response.status_code == 429:
                retry_after = int(response.headers.get("Retry-After", "30"))
                logging.warning("Rate limited by Liquipedia; waiting %ss", retry_after + 10)
                time.sleep(retry_after + 10)
                continue
            if response.status_code == 503:
                time.sleep(30)
                continue
            if response.status_code >= 500:
                time.sleep(15)
                continue

            logging.warning("Unexpected status %s for %s", response.status_code, url)
            return None

        self.summary.errors += 1
        return None

    def request_delay_seconds(self) -> float:
        jitter = random.triangular(self.jitter_min, self.jitter_max, self.jitter_mode)
        return max(2.0, self.rate_limit + jitter)

    def discover_tournaments(self) -> list[Tournament]:
        tournaments: dict[str, Tournament] = {}
        for seed_path in discovery_seed_paths(self.config):
            html = self.fetch(seed_path)
            if not html:
                continue
            soup = make_soup(html)
            seed_tier = infer_tier(seed_path)
            seed_year = infer_date_hint(seed_path)
            if seed_tier in {"s", "a"} and not seed_year:
                discovered = tournaments_from_tier_page_before_2021(soup, seed_tier)
            elif seed_tier:
                discovered = tournaments_from_tournament_table(soup, seed_tier, seed_year)
            else:
                discovered = tournaments_from_index(soup, seed_tier, seed_year)
            for tournament in discovered:
                if not seed_tier and tournament.tier in {"s", "a"}:
                    continue
                if self.in_scope(tournament):
                    tournaments[tournament.path] = tournament
                else:
                    self.summary.rows_skipped_scope += 1

        items = sorted(tournaments.values(), key=lambda item: item.path)
        if self.mode == "incremental":
            processed = set(self.state.get("processed_tournament_urls", []))
            items = [item for item in items if item.path not in processed]
        logging.info("Discovered %s in-scope tournaments", len(items))
        return items

    def in_scope(self, tournament: Tournament) -> bool:
        title = tournament.title.lower()
        date_text = tournament.date_hint or title
        event_date = parse_any_date(date_text)
        if not event_date:
            return True
        cutoffs = self.config["backfill"]
        if "overwatch league" in title or "/overwatch_league/" in tournament.path.lower():
            return event_date >= datetime.fromisoformat(cutoffs["owl_from"])
        if "contenders" in title:
            return event_date >= datetime.fromisoformat(cutoffs["contenders_from"])
        if tournament.tier in {"s", "a"}:
            return event_date >= datetime(2022, 1, 1)
        return event_date >= datetime.fromisoformat(cutoffs["open_tournaments_from"])

    def parse_tournament(self, tournament: Tournament) -> list[ParsedMatch]:
        matches: list[ParsedMatch] = []
        for path in [tournament.path, f"{tournament.path.rstrip('/')}/Bracket"]:
            html = self.fetch(path)
            if not html:
                continue
            page_matches = self.parse_tournament_html(html, tournament, path)
            if not page_matches and path.endswith("/Bracket"):
                logging.info("No bracket matches parsed for %s", tournament.path)
            matches.extend(page_matches)
        return matches

    def parse_tournament_html(self, html: str, tournament: Tournament, source_path: str) -> list[ParsedMatch]:
        soup = make_soup(html)
        page_title = extract_page_title(soup)
        date_hint, date_hint_source = extract_tournament_date_hint(soup, tournament, source_path)
        page_tournament = Tournament(
            title=page_title or tournament.title,
            path=tournament.path,
            tier=tournament.tier or infer_tier(soup.get_text(" ", strip=True)),
            date_hint=date_hint,
            date_hint_source=date_hint_source,
        )
        roster_map = extract_participant_rosters(soup)
        matches = self.parse_match_tables(soup, page_tournament, source_path, roster_map)
        if soup.select(".brkts-match") and not matches:
            logging.info(
                "Parsed 0 matches from %s despite %s bracket nodes; date_hint=%r rosters=%s",
                source_path,
                len(soup.select(".brkts-match")),
                date_hint,
                len(roster_map),
            )
        return matches

    def parse_direct_pages(self) -> list[ParsedMatch]:
        matches: list[ParsedMatch] = []
        for page in self.page_paths:
            path = liquipedia_path(page, self.base_url)
            html = self.fetch(path)
            if not html:
                continue
            logging.info("Processing direct Liquipedia page %s", path)
            tournament = Tournament(title=title_from_path(path), path=path)
            page_matches = self.parse_tournament_html(html, tournament, path)
            self.summary.tournaments_processed += 1
            self.summary.matches_found += len(page_matches)
            matches.extend(page_matches)

        for html_file in self.html_files:
            if not html_file.exists():
                logging.warning("HTML file not found: %s", html_file)
                self.summary.errors += 1
                continue
            logging.info("Processing local Liquipedia HTML file %s", html_file)
            html = html_file.read_text(encoding="utf-8", errors="replace")
            source_path = str(html_file)
            tournament = Tournament(title=html_file.stem, path=source_path)
            page_matches = self.parse_tournament_html(html, tournament, source_path)
            self.summary.tournaments_processed += 1
            self.summary.matches_found += len(page_matches)
            matches.extend(page_matches)
        return matches

    def parse_match_tables(self, soup: BeautifulSoup, tournament: Tournament, source_path: str, roster_map: dict[str, list[str]] | None = None) -> list[ParsedMatch]:
        roster_map = roster_map or {}
        matches: list[ParsedMatch] = self.parse_bracket_matches(soup, tournament, source_path, roster_map)
        for table in soup.select("table.wikitable, table.matchlist, div.bracket"):
            stage = nearest_heading(table) or tournament.tier or "Unknown"
            rows = table.select("tr") if table.name == "table" else table.select(".bracket-game, .bracket-popup-body")
            for row in rows:
                try:
                    parsed = self.parse_match_row(row, tournament, stage, source_path, roster_map)
                except Exception as exc:
                    self.summary.parse_warnings += 1
                    logging.warning("Failed parsing row on %s: %s", source_path, exc)
                    continue
                if parsed and not match_has_no_roster(parsed):
                    matches.append(parsed)
        return unique_matches(matches)

    def parse_bracket_matches(self, soup: BeautifulSoup, tournament: Tournament, source_path: str, roster_map: dict[str, list[str]]) -> list[ParsedMatch]:
        matches: list[ParsedMatch] = []
        for node in soup.select(".brkts-match"):
            try:
                parsed = self.parse_bracket_match(node, tournament, source_path, roster_map)
            except Exception as exc:
                self.summary.parse_warnings += 1
                logging.warning("Failed parsing bracket match on %s: %s", source_path, exc)
                continue
            if parsed and not match_has_no_roster(parsed):
                matches.append(parsed)
        return matches

    def parse_bracket_match(self, node: Tag, tournament: Tournament, source_path: str, roster_map: dict[str, list[str]]) -> ParsedMatch | None:
        opponent_nodes = node.select(":scope > .brkts-opponent-entry")
        if len(opponent_nodes) < 2:
            return None

        team_a = bracket_team_name(opponent_nodes[0])
        team_b = bracket_team_name(opponent_nodes[1])
        if not team_a or not team_b:
            return None

        score_a = bracket_score(opponent_nodes[0])
        score_b = bracket_score(opponent_nodes[1])
        if score_a is None or score_b is None:
            return None

        popup = node.select_one(".brkts-popup")
        if popup:
            popup_teams = popup_team_names(popup)
            if len(popup_teams) >= 2:
                team_a, team_b = popup_teams[:2]

            popup_scores = popup_score_values(popup)
            if len(popup_scores) >= 2:
                score_a, score_b = popup_scores[:2]

        timestamp_node = node.select_one("[data-timestamp]")
        timestamp_value = str(timestamp_node.get("data-timestamp") or "") if timestamp_node else ""
        parsed_date = parse_any_date(timestamp_value)
        match_date = timestamp_value
        if not parsed_date:
            match_date = tournament.date_hint
            parsed_date = parse_any_date(str(match_date or ""))
        if not parsed_date:
            self.summary.parse_warnings += 1
            logging.warning("Could not extract bracket date for %s vs %s on %s", team_a, team_b, source_path)
            return None

        best_of = bracket_best_of(node) or infer_best_of(score_a, score_b)
        stage = nearest_heading(node) or "Bracket"
        match = ParsedMatch(
            match_date=parsed_date.isoformat(),
            event_name=tournament.title,
            stage=stage,
            team_a=team_a,
            team_b=team_b,
            score_a=score_a,
            score_b=score_b,
            best_of=best_of,
            source_url=urljoin(self.base_url, source_path),
        )
        match.notes.append(f"liquipedia_source: {match.source_url}")
        match.notes.append(f"teams: {match.team_a} vs {match.team_b}")
        if not parse_any_date(timestamp_value) and tournament.date_hint_source and tournament.date_hint_source != "start date":
            match.notes.append(f"date_fallback: {tournament.date_hint_source} used because this bracket match has no timestamp")
        self.apply_rosters(match, node, parsed_date, roster_map)
        return match

    def parse_match_row(self, row: Tag, tournament: Tournament, stage: str, source_path: str, roster_map: dict[str, list[str]]) -> ParsedMatch | None:
        text = clean(row.get_text(" ", strip=True))
        if not text or not re.search(r"\b\d+\s*[-:]\s*\d+\b", text):
            return None

        score_match = re.search(r"\b(\d+)\s*[-:]\s*(\d+)\b", text)
        if not score_match:
            return None
        score_a, score_b = int(score_match.group(1)), int(score_match.group(2))
        if score_a == score_b:
            winning_side = "draw"
        else:
            winning_side = "a" if score_a > score_b else "b"

        names = extract_team_names(row)
        if len(names) < 2:
            self.summary.parse_warnings += 1
            logging.warning("Could not extract two teams from %s row: %s", source_path, text[:240])
            return None

        row_timestamp = extract_timestamp(row)
        match_date = row_timestamp or tournament.date_hint
        parsed_date = parse_any_date(match_date)
        if not parsed_date:
            self.summary.parse_warnings += 1
            logging.warning("Could not extract date for %s row: %s", source_path, text[:240])
            return None

        best_of = infer_best_of(score_a, score_b)
        match = ParsedMatch(
            match_date=parsed_date.isoformat(),
            event_name=tournament.title,
            stage=stage,
            team_a=names[0],
            team_b=names[1],
            score_a=score_a,
            score_b=score_b,
            best_of=best_of,
            source_url=urljoin(self.base_url, source_path),
        )
        match.notes.append(f"liquipedia_source: {match.source_url}")
        match.notes.append(f"teams: {match.team_a} vs {match.team_b}")
        if not row_timestamp and tournament.date_hint_source and tournament.date_hint_source != "start date":
            match.notes.append(f"date_fallback: {tournament.date_hint_source} used because this match row has no timestamp")
        self.apply_rosters(match, row, parsed_date, roster_map)
        return match

    def apply_rosters(self, match: ParsedMatch, row: Tag, match_date: datetime, roster_map: dict[str, list[str]]) -> None:
        exact_a, exact_b = extract_inline_players(row)
        if len(exact_a) >= 5 and len(exact_b) >= 5:
            match.players_a = exact_a
            match.players_b = exact_b
            match.roster_confidence = "exact"
            self.summary.exact += 1
            return

        roster_a = roster_for_team(roster_map, match.team_a)
        roster_b = roster_for_team(roster_map, match.team_b)
        match.players_a = roster_a
        match.players_b = roster_b

        if len(match.players_a) >= 5 and len(match.players_b) >= 5:
            match.roster_confidence = "inferred"
            match.notes.append(f"roster_inferred: tournament page roster used for {match_date.date()}; please verify")
            self.summary.inferred += 1
        elif match.players_a or match.players_b:
            match.roster_confidence = "partial"
            match.notes.append(f"roster_partial: side A {len(match.players_a)}/5, side B {len(match.players_b)}/5; manual fill required")
            self.summary.partial += 1
        else:
            match.roster_confidence = "missing"
            match.notes.append("roster_missing: no roster data found; manual entry required")
            self.summary.missing += 1

    def alias_map(self, workbook) -> dict[str, str]:
        ensure_sheet(workbook, "Player Alias Map", ALIAS_HEADERS)
        sheet = workbook["Player Alias Map"]
        aliases: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            canonical, alias, _source, approved = (list(row) + [None] * 4)[:4]
            if str(approved or "yes").strip().lower() not in {"yes", "y", "true", "1"}:
                continue
            if canonical:
                aliases[normalize_name(str(canonical))] = clean(str(canonical))
            if alias and canonical:
                aliases[normalize_name(str(alias))] = clean(str(canonical))
        return aliases

    def existing_rows(self, workbook) -> list[dict[str, Any]]:
        ensure_sheet(workbook, "Match Entry", MATCH_HEADERS)
        sheet = workbook["Match Entry"]
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for idx, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(values):
                continue
            data = {str(headers[pos]): value for pos, value in enumerate(values) if pos < len(headers)}
            data["_row"] = idx
            rows.append(data)
        return rows

    def write_matches(self, matches: list[ParsedMatch]) -> tuple[int, int]:
        workbook = open_or_create_workbook(self.workbook_path)
        ensure_sheet(workbook, "Match Entry", MATCH_HEADERS)
        ensure_sheet(workbook, "Player Alias Map", ALIAS_HEADERS)
        ensure_sheet(workbook, "Import Log", IMPORT_LOG_HEADERS)

        aliases = self.alias_map(workbook)
        existing = self.existing_rows(workbook)
        sheet = workbook["Match Entry"]
        headers = [cell.value for cell in sheet[1]]
        written = 0
        skipped = 0

        for match in matches:
            row = self.build_workbook_row(match, aliases)
            ensure_sheet(workbook, "Match Entry", match_headers_for_row(row))
            headers = [cell.value for cell in sheet[1]]
            duplicate = find_duplicate(row, existing)
            if duplicate:
                skipped += 1
                self.summary.rows_skipped_dup += 1
                logging.info("duplicate_suspected: matches row %s for %s", duplicate, match.event_name)
                continue

            if self.mode == "dry-run":
                logging.info("DRY RUN write: %s %s %s-%s", row["event_name"], row["match_date"], row["score_a"], row["score_b"])
            else:
                sheet.append([row.get(header, "") for header in headers])
                existing.append(row)
            written += 1
            if row["status"] == "ready":
                self.summary.rows_ready += 1
            else:
                self.summary.rows_draft += 1

        if self.mode != "dry-run":
            log_sheet = workbook["Import Log"]
            log_sheet.append([datetime.now().isoformat(), len(matches), written, skipped, "Liquipedia backfill bot"])
            workbook.save(self.workbook_path)
        return written, skipped

    def build_workbook_row(self, match: ParsedMatch, aliases: dict[str, str]) -> dict[str, Any]:
        players_a, unresolved_a = resolve_names(match.players_a, aliases)
        players_b, unresolved_b = resolve_names(match.players_b, aliases)
        unresolved = unresolved_a + unresolved_b
        self.summary.unresolved_players += len(unresolved)

        notes = list(match.notes)
        for name in unresolved:
            notes.append(f"unresolved_player: '{name}' not in alias map")

        status = "ready"
        if match.roster_confidence != "exact" or unresolved or len(players_a) < 5 or len(players_b) < 5:
            status = "draft"

        winning_side = "draw"
        if match.score_a > match.score_b:
            winning_side = "a"
        elif match.score_b > match.score_a:
            winning_side = "b"

        row = {
            "match_date": match.match_date,
            "event_name": match.event_name,
            "stage": match.stage,
            "score_a": match.score_a,
            "score_b": match.score_b,
            "best_of": f"BO{match.best_of}",
            "winning_side": winning_side,
            "notes": "; ".join(notes),
            "status": status,
            "import_error": "",
            "row_hash": "",
            "roster_confidence": match.roster_confidence,
        }
        for index, player in enumerate(players_a, start=1):
            row[f"player_a{index}"] = player
        for index, player in enumerate(players_b, start=1):
            row[f"player_b{index}"] = player
        return row

    def run(self) -> int:
        started = time.monotonic()
        logging.info("Liquipedia bot started in %s mode", self.mode)
        processed: list[str] = []
        all_matches: list[ParsedMatch] = self.parse_direct_pages()

        if not self.page_paths and not self.html_files:
            tournaments = self.discover_tournaments()
            for tournament in tournaments:
                logging.info("Processing %s", tournament.path)
                matches = self.parse_tournament(tournament)
                self.summary.tournaments_processed += 1
                self.summary.matches_found += len(matches)
                all_matches.extend(matches)
                processed.append(tournament.path)

        written, skipped = self.write_matches(all_matches)
        self.save_state(written, skipped, processed)
        duration = timedelta(seconds=int(time.monotonic() - started))
        self.log_summary(duration)
        return 0

    def log_summary(self, duration: timedelta) -> None:
        logging.info("[BACKFILL COMPLETE]")
        logging.info("Tournaments processed  : %s", self.summary.tournaments_processed)
        logging.info("Matches found          : %s", self.summary.matches_found)
        logging.info("  -> exact roster       : %s", self.summary.exact)
        logging.info("  -> inferred roster    : %s", self.summary.inferred)
        logging.info("  -> partial roster     : %s", self.summary.partial)
        logging.info("  -> missing roster     : %s", self.summary.missing)
        logging.info("Rows written (ready)   : %s", self.summary.rows_ready)
        logging.info("Rows written (draft)   : %s", self.summary.rows_draft)
        logging.info("Rows skipped (dup)     : %s", self.summary.rows_skipped_dup)
        logging.info("Rows skipped (scope)   : %s", self.summary.rows_skipped_scope)
        logging.info("Pages served from cache: %s", self.summary.cache_hits)
        logging.info("Unresolved players     : %s", self.summary.unresolved_players)
        logging.info("Parse warnings         : %s", self.summary.parse_warnings)
        logging.info("Errors                 : %s", self.summary.errors)
        logging.info("Duration               : %s", duration)


def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def discovery_seed_paths(config: dict[str, Any]) -> list[str]:
    seeds: list[str] = []
    current_year = min(datetime.now().year, BACKFILL_END_YEAR)
    seeds.extend([f"/overwatch/{tier}-Tier_Tournaments" for tier in ("S", "A")])
    for tier in ("B", "C", "D"):
        for year in range(2023, current_year + 1):
            seeds.append(f"/overwatch/{tier}-Tier_Tournaments/{year}")
    seeds.extend([f"/overwatch/{tier}-Tier_Tournaments" for tier in ("B", "C", "D")])
    return list(dict.fromkeys(seeds))


def tournaments_from_index(soup: BeautifulSoup, seed_tier: str = "", seed_year: str = "") -> list[Tournament]:
    tournaments: dict[str, Tournament] = {}
    for anchor in soup.find_all("a", href=True):
        href = clean(anchor.get("href", ""))
        title = clean(anchor.get("title") or anchor.get_text(" ", strip=True))
        if not title or not is_tournament_path(href):
            continue
        row_text = clean(parent_text(anchor))
        tier = infer_tier(row_text + " " + title) or seed_tier
        date_hint = infer_date_hint(row_text)
        if not seed_tier and (not tier or not date_hint):
            continue
        if not date_hint and seed_year != "2023":
            date_hint = seed_year
        tournaments[href] = Tournament(title=title, path=href, tier=tier, date_hint=date_hint)
    return list(tournaments.values())


def tournaments_from_tier_page_before_2021(soup: BeautifulSoup, seed_tier: str) -> list[Tournament]:
    return tournaments_from_tournament_table(soup, seed_tier, stop_at_year=2021)


def tournaments_from_tournament_table(soup: BeautifulSoup, seed_tier: str = "", seed_year: str = "", stop_at_year: int | None = None) -> list[Tournament]:
    tournaments: dict[str, Tournament] = {}
    for row in soup.select("tr"):
        row_text = clean(row.get_text(" ", strip=True))
        if not row_text:
            continue
        row_date = parse_tournament_row_date(row_text)
        if stop_at_year and row_date and row_date.year <= stop_at_year:
            break
        anchor = row.select_one(".column__tournament a[href^='/overwatch/'][title]") or row.select_one("td.column__tournament a[href^='/overwatch/'][title]")
        if not anchor:
            continue
        href = clean(anchor.get("href", ""))
        title = clean(anchor.get("title") or anchor.get_text(" ", strip=True))
        if not title or not is_tournament_path(href):
            continue
        date_hint = row_date.date().isoformat() if row_date else infer_date_hint(row_text) or seed_year
        tournaments[href] = Tournament(title=title, path=href, tier=seed_tier, date_hint=date_hint)
    return list(tournaments.values())


def parse_tournament_row_date(text: str) -> datetime | None:
    text = clean(text)
    range_match = re.search(r"\b([A-Z][a-z]{2,8})\.?\s+\d{1,2}\s*[–-]\s*(?:[A-Z][a-z]{2,8}\.?\s+)?\d{1,2},\s*(20\d{2})\b", text)
    if range_match:
        month = range_match.group(1).rstrip(".")
        year = range_match.group(2)
        return parse_any_date(f"{month} 1, {year}")
    full_date = re.search(r"\b[A-Z][a-z]{2,8}\.?\s+\d{1,2},\s*20\d{2}\b", text)
    if full_date:
        return parse_any_date(full_date.group(0).replace(".", ""))
    year = re.search(r"\b20\d{2}\b", text)
    return parse_any_date(year.group(0)) if year else None


def make_soup(html: str) -> BeautifulSoup:
    try:
        return BeautifulSoup(html, HTML_PARSER)
    except Exception:
        logging.warning("lxml parser unavailable; falling back to html.parser")
        return BeautifulSoup(html, "html.parser")


def normalize_name(text: str) -> str:
    return clean(text).casefold()


def normalize_team(text: str) -> str:
    text = re.sub(r"\s+\(page does not exist\)$", "", clean(text))
    text = re.sub(r"\s+\([^)]*team\)$", "", text, flags=re.IGNORECASE)
    return text.casefold()


def liquipedia_path(page: str, base_url: str) -> str:
    parsed = urlparse(page)
    if parsed.scheme and parsed.netloc:
        path = parsed.path
    else:
        path = page
    if not path.startswith("/"):
        path = f"/overwatch/{path}"
    return path


def title_from_path(path: str) -> str:
    return clean(path.rstrip("/").split("/")[-1].replace("_", " ")) or path


def extract_page_title(soup: BeautifulSoup) -> str:
    heading = soup.select_one("#firstHeading")
    return clean(heading.get_text(" ", strip=True)) if heading else ""


def extract_infobox_value(soup: BeautifulSoup, label: str) -> str:
    wanted = label.strip().rstrip(":").casefold()
    for description in soup.select(".infobox-description"):
        text = clean(description.get_text(" ", strip=True)).rstrip(":").casefold()
        if text != wanted:
            continue
        value = description.find_next_sibling()
        if value:
            return clean(value.get_text(" ", strip=True))
    return ""


def extract_tournament_date_hint(soup: BeautifulSoup, tournament: Tournament, source_path: str) -> tuple[str, str]:
    for label, source in (
        ("Start Date", "start date"),
        ("Date", "date field"),
        ("Dates", "dates field"),
        ("Start", "start field"),
    ):
        value = extract_infobox_value(soup, label)
        if parse_any_date(value):
            return value, source

    if parse_any_date(tournament.date_hint):
        return tournament.date_hint, "discovery date"

    page_title = extract_page_title(soup)
    title_hint = infer_date_hint(f"{page_title} {tournament.title}")
    if parse_any_date(title_hint):
        return title_hint, "page title year"

    path_hint = infer_date_hint_from_path(source_path or tournament.path)
    if parse_any_date(path_hint):
        return path_hint, "page path year"

    return tournament.date_hint, "missing"


def extract_participant_rosters(soup: BeautifulSoup) -> dict[str, list[str]]:
    rosters: dict[str, list[str]] = {}
    for card in soup.select(".team-participant-card"):
        team = participant_card_team(card)
        players = participant_card_players(card)
        if team and players:
            rosters[normalize_team(team)] = players
            rosters[normalize_name(team)] = players
    for card in soup.select(".teamcard"):
        team = teamcard_team(card)
        players = teamcard_players(card)
        if team and players:
            rosters[normalize_team(team)] = players
            rosters[normalize_name(team)] = players
    return rosters


def participant_card_team(card: Tag) -> str:
    header = card.select_one(".team-participant-card__header") or card
    link = header.select_one(".block-team a[title]")
    if link and link.get("title"):
        title = re.sub(r"\s+\(page does not exist\)$", "", str(link["title"]))
        return clean(title)
    name = header.select_one(".block-team .name")
    return clean(name.get_text(" ", strip=True)) if name else ""


def participant_card_players(card: Tag) -> list[str]:
    players: list[str] = []
    selector = ".toggle-area-content-active[data-toggle-area-content='1'] .team-participant-card__member-name .name"
    members = card.select(selector) or card.select(".team-participant-card__member-name .name")
    for member in members:
        name = clean(member.get_text(" ", strip=True))
        if name and name not in players:
            players.append(name)
    return players


def teamcard_team(card: Tag) -> str:
    heading = card.find("center")
    if heading:
        link = heading.select_one("a[title]")
        if link:
            return clean(str(link.get("title") or link.get_text(" ", strip=True)))
        text = clean(heading.get_text(" ", strip=True))
        if text:
            return text
    link = card.select_one(".center a[title]")
    return clean(str(link.get("title") or link.get_text(" ", strip=True))) if link else ""


def teamcard_players(card: Tag) -> list[str]:
    players: list[str] = []
    main_tables = card.select("table[data-toggle-area-content='1']")
    tables = main_tables or card.select("table.wikitable")
    for table in tables:
        for row in table.select("tr"):
            if "teamcard-bg-dnp" in row.get("class", []):
                continue
            text = clean(row.get_text(" ", strip=True))
            if re.search(r"\b(DNP|Did not play)\b", text, re.IGNORECASE):
                continue
            role = row.select_one("th img[alt]")
            if role and role.get("alt", "").lower() in {"coach", "assistant coach", "manager"}:
                continue
            name = player_name_from_roster_cell(row)
            if name and name not in players:
                players.append(name)
    return players


def player_name_from_roster_cell(row: Tag) -> str:
    cells = row.select("td")
    search_root = cells[-1] if cells else row
    anchors = search_root.select("a[href^='/overwatch/']")
    for anchor in reversed(anchors):
        href = anchor.get("href", "")
        name = clean(anchor.get_text(" ", strip=True))
        title = clean(anchor.get("title") or name)
        if not name:
            continue
        if "/Category:" in href or "/File:" in href:
            continue
        if anchor.find_parent(class_="team-template-team-part") or anchor.find_parent(class_="flag") or anchor.find_parent(class_="team-template-image-icon"):
            continue
        if re.search(r"(Coach|Manager|Team|Tournament|League)$", title, re.IGNORECASE):
            continue
        return name
    return ""


def roster_for_team(rosters: dict[str, list[str]], team_name: str) -> list[str]:
    candidates = [normalize_team(team_name), normalize_name(team_name)]
    for key in candidates:
        if key in rosters:
            return rosters[key]
    for key, players in rosters.items():
        if key in candidates[0] or candidates[0] in key:
            return players
    return []


def parent_text(anchor: Tag) -> str:
    parent = anchor.find_parent("tr") or anchor.find_parent("li") or anchor.parent
    return parent.get_text(" ", strip=True) if parent else ""


def is_tournament_path(href: str) -> bool:
    if not href.startswith("/overwatch/") or ":" in href:
        return False
    blocked = (
        "/Portal",
        "/Liquipedia",
        "/Special:",
        "/File:",
        "/Category:",
        "/Main_Page",
        "/index.php",
        "_Tier_Tournaments",
        "/Played_Matches",
        "/Results",
    )
    return not any(part.lower() in href.lower() for part in blocked)


def infer_tier(text: str) -> str:
    match = re.search(r"\b([SABCD])[- _]?Tier(?:\b|_)", text, re.IGNORECASE)
    return match.group(1).lower() if match else ""


def infer_date_hint(text: str) -> str:
    match = re.search(r"\b(20\d{2}[-/]\d{1,2}[-/]\d{1,2}|[A-Z][a-z]+ \d{1,2}, 20\d{2}|20\d{2})\b", text)
    return match.group(1) if match else ""


def infer_date_hint_from_path(path: str) -> str:
    decoded = unquote(path).replace("_", " ")
    full_year = re.search(r"\b(20\d{2})(?:\s*[-/]\s*\d{2})?\b", decoded)
    if full_year:
        return full_year.group(1)

    two_digit_segment = re.search(r"/(\d{2})(?:/|$)", decoded)
    if not two_digit_segment:
        return ""
    year = int(two_digit_segment.group(1))
    current_short_year = BACKFILL_END_YEAR % 100
    if 20 <= year <= current_short_year:
        return f"20{year:02d}"
    return ""


def parse_any_date(text: str) -> datetime | None:
    if not text:
        return None
    text = clean(str(text))
    if text.isdigit() and len(text) == 10:
        return datetime.fromtimestamp(int(text), tz=timezone.utc).replace(tzinfo=None)
    patterns = [
        r"20\d{2}-\d{1,2}-\d{1,2}T\d{1,2}:\d{2}:\d{2}",
        r"20\d{2}-\d{1,2}-\d{1,2} \d{1,2}:\d{2}:\d{2}",
        r"20\d{2}-\d{1,2}-\d{1,2}",
        r"20\d{2}/\d{1,2}/\d{1,2}",
        r"[A-Z][a-z]+ \d{1,2}, 20\d{2}",
        r"[A-Z][a-z]{2} \d{1,2}, 20\d{2}",
        r"20\d{2}",
    ]
    formats = ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%B %d, %Y", "%b %d, %Y", "%Y")
    for pattern, fmt in zip(patterns, formats):
        found = re.search(pattern, text)
        if not found:
            continue
        try:
            parsed = datetime.strptime(found.group(0), fmt)
            return parsed if fmt != "%Y" else datetime(parsed.year, 1, 1)
        except ValueError:
            continue
    return None


def extract_timestamp(row: Tag) -> str:
    stamped = row.select_one("[data-timestamp]")
    if stamped and stamped.get("data-timestamp"):
        return str(stamped["data-timestamp"])
    return infer_date_hint(row.get_text(" ", strip=True))


def extract_team_names(row: Tag) -> list[str]:
    candidates: list[str] = []
    for selector in [".team-template-text a", ".team-template-team-standard a", "span[title]", "a[title]"]:
        for node in row.select(selector):
            text = clean(node.get("title") or node.get_text(" ", strip=True))
            if text and text not in candidates and not text.isdigit():
                candidates.append(text)
    if len(candidates) >= 2:
        return candidates[:2]

    cells = [clean(cell.get_text(" ", strip=True)) for cell in row.select("td")]
    cells = [cell for cell in cells if cell and not re.search(r"^\d+\s*[-:]\s*\d+$", cell)]
    return cells[:2]


def bracket_team_name(opponent: Tag) -> str:
    label = clean(opponent.get("aria-label") or "")
    if label:
        return label
    name = opponent.select_one(".name.hidden-xs") or opponent.select_one(".name")
    return clean(name.get_text(" ", strip=True)) if name else ""


def bracket_score(opponent: Tag) -> int | None:
    score = opponent.select_one(".brkts-opponent-score-inner")
    if not score:
        return None
    text = clean(score.get_text(" ", strip=True))
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def popup_team_names(popup: Tag) -> list[str]:
    names: list[str] = []
    for opponent in popup.select(".match-info-header-opponent"):
        link = opponent.select_one(".block-team a[title]")
        raw = clean(link.get("title") if link else "")
        raw = re.sub(r"\s+\(page does not exist\)$", "", raw)
        if raw and raw not in names:
            names.append(raw)
            continue
        name = opponent.select_one(".name")
        text = clean(name.get_text(" ", strip=True)) if name else ""
        if text and text not in names:
            names.append(text)
    return names


def popup_score_values(popup: Tag) -> list[int]:
    values: list[int] = []
    for score in popup.select(".match-info-header-scoreholder-score"):
        text = clean(score.get_text(" ", strip=True))
        if text.isdigit():
            values.append(int(text))
    return values


def bracket_best_of(node: Tag) -> int | None:
    lower = node.select_one(".match-info-header-scoreholder-lower")
    if not lower:
        return None
    match = re.search(r"Bo\s*(\d+)", lower.get_text(" ", strip=True), re.IGNORECASE)
    return int(match.group(1)) if match else None


def extract_inline_players(row: Tag) -> tuple[list[str], list[str]]:
    if not row.select(".match-info-player, .match-info-players, .match-info-roster"):
        return [], []
    player_links = []
    for anchor in row.select("a[href^='/overwatch/']"):
        parent_row = anchor.find_parent("tr")
        if parent_row and "teamcard-bg-dnp" in parent_row.get("class", []):
            continue
        href = anchor.get("href", "")
        text = clean(anchor.get("title") or anchor.get_text(" ", strip=True))
        if text and "/overwatch/" in href and ":" not in href and text not in player_links:
            if not re.search(r"(League|Contenders|Cup|Team|Championship|Tournament)", text, re.IGNORECASE):
                player_links.append(text)
    if len(player_links) >= 10:
        midpoint = len(player_links) // 2
        return player_links[:midpoint], player_links[midpoint:]
    return [], []


def parse_roster_history(soup: BeautifulSoup, match_date: datetime) -> list[str]:
    players: list[str] = []
    for table in soup.select("table.wikitable"):
        header = clean(table.get_text(" ", strip=True)).lower()
        if "join" not in header and "leave" not in header and "roster" not in header:
            continue
        for row in table.select("tr"):
            text = clean(row.get_text(" ", strip=True))
            name = first_player_name(row)
            if not name:
                continue
            dates = re.findall(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}|[A-Z][a-z]+ \d{1,2}, 20\d{2}", text)
            join = parse_any_date(dates[0]) if dates else None
            leave = parse_any_date(dates[1]) if len(dates) > 1 else None
            if (not join or join <= match_date) and (not leave or leave >= match_date):
                if name not in players:
                    players.append(name)
    return players


def parse_current_roster(soup: BeautifulSoup) -> list[str]:
    players: list[str] = []
    for node in soup.select(".table-responsive tbody tr, table.wikitable tr"):
        text = clean(node.get_text(" ", strip=True)).lower()
        if any(word in text for word in ("active", "player", "roster")):
            name = first_player_name(node)
            if name and name not in players:
                players.append(name)
    return players


def first_player_name(row: Tag) -> str:
    for anchor in row.select("a[href^='/overwatch/']"):
        name = clean(anchor.get("title") or anchor.get_text(" ", strip=True))
        if name and not re.search(r"(Team|Coach|Manager|Tournament|League)", name, re.IGNORECASE):
            return name
    return ""


def nearest_heading(node: Tag) -> str:
    cursor = node
    while cursor:
        cursor = cursor.find_previous(["h2", "h3", "h4"])
        if not cursor:
            return ""
        headline = clean(cursor.get_text(" ", strip=True)).replace("[ edit ]", "")
        if headline:
            return headline
    return ""


def infer_best_of(score_a: int, score_b: int) -> int:
    wins_needed = max(score_a, score_b)
    return max(1, wins_needed * 2 - 1)


def unique_matches(matches: list[ParsedMatch]) -> list[ParsedMatch]:
    seen: set[tuple[str, str, str, int, int]] = set()
    unique: list[ParsedMatch] = []
    for match in matches:
        key = (match.match_date, match.team_a, match.team_b, match.score_a, match.score_b)
        if key not in seen:
            unique.append(match)
            seen.add(key)
    return unique


def match_has_no_roster(match: ParsedMatch) -> bool:
    return len(match.players_a) == 0 or len(match.players_b) == 0


def resolve_names(names: list[str], aliases: dict[str, str]) -> tuple[list[str], list[str]]:
    resolved: list[str] = []
    unresolved: list[str] = []
    for name in names:
        canonical = aliases.get(normalize_name(name))
        if canonical:
            resolved.append(canonical)
        else:
            resolved.append(clean(name))
            unresolved.append(clean(name))
    return resolved, unresolved


def open_or_create_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    workbook.active.title = "Match Entry"
    ensure_sheet(workbook, "Match Entry", MATCH_HEADERS)
    ensure_sheet(workbook, "Player Alias Map", ALIAS_HEADERS)
    ensure_sheet(workbook, "Import Log", IMPORT_LOG_HEADERS)
    return workbook


def ensure_sheet(workbook, name: str, headers: list[str]) -> None:
    sheet = workbook[name] if name in workbook.sheetnames else workbook.create_sheet(name)
    if sheet.max_row == 1 and not any(cell.value for cell in sheet[1]):
        for idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=idx, value=header)
        return
    existing = [cell.value for cell in sheet[1]]
    for header in headers:
        if header not in existing:
            sheet.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)


def match_headers_for_row(row: dict[str, Any]) -> list[str]:
    max_a = max([int(key.replace("player_a", "")) for key in row if re.match(r"^player_a\d+$", key)] or [11])
    max_b = max([int(key.replace("player_b", "")) for key in row if re.match(r"^player_b\d+$", key)] or [11])
    return [
        *MATCH_PREFIX_HEADERS,
        *[f"player_a{i}" for i in range(1, max(11, max_a) + 1)],
        *[f"player_b{i}" for i in range(1, max(11, max_b) + 1)],
        *MATCH_SUFFIX_HEADERS,
    ]


def find_duplicate(row: dict[str, Any], existing: list[dict[str, Any]]) -> int | None:
    row_date = parse_any_date(str(row.get("match_date", "")))
    row_players = player_set(row)
    row_score = (str(row.get("score_a", "")).strip(), str(row.get("score_b", "")).strip())
    for candidate in existing:
        if clean(str(candidate.get("event_name", ""))) != clean(str(row.get("event_name", ""))):
            continue
        candidate_score = (str(candidate.get("score_a", "")).strip(), str(candidate.get("score_b", "")).strip())
        if row_score != candidate_score:
            continue
        candidate_date = parse_any_date(str(candidate.get("match_date", "")))
        if row_date and candidate_date and abs((row_date - candidate_date).total_seconds()) > 3600:
            continue
        overlap = len(row_players.intersection(player_set(candidate)))
        if overlap >= 8:
            return int(candidate.get("_row", 0))
    return None


def player_set(row: dict[str, Any]) -> set[str]:
    values = []
    for key, value in row.items():
        if re.match(r"^player_[ab]\d+$", str(key)) and value:
            values.append(normalize_name(str(value)))
    return set(values)


def main() -> int:
    parser = argparse.ArgumentParser(description="Liquipedia historical OWCS workbook backfill bot")
    parser.add_argument("--mode", choices=["backfill", "incremental", "dry-run"], required=True)
    parser.add_argument(
        "--page",
        action="append",
        default=[],
        help="Parse a specific Liquipedia tournament page path or URL instead of crawling discovery pages.",
    )
    parser.add_argument(
        "--html-file",
        action="append",
        default=[],
        help="Parse a local saved Liquipedia HTML page, useful for validating parser changes against Example.html.",
    )
    parser.add_argument(
        "--terminal-doc",
        default="",
        help="Write the same terminal-style logging output to this troubleshooting file.",
    )
    args = parser.parse_args()
    return LiquipediaBot(args.mode, pages=args.page, html_files=args.html_file, terminal_doc=args.terminal_doc).run()


if __name__ == "__main__":
    raise SystemExit(main())
