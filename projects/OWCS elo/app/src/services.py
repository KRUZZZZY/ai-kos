from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import load_config, project_path
from .database import get_db
from .elo import rebuild_ratings

DEFAULT_PLAYER_COLUMNS = [f"player_a{i}" for i in range(1, 6)] + [f"player_b{i}" for i in range(1, 6)]
PLAYER_COLUMN_RE = re.compile(r"^player_([ab])(\d+)$")
FIXED_HASH_COLUMNS = [
    "match_date",
    "event_name",
    "stage",
    "score_a",
    "score_b",
    "best_of",
    "winning_side",
]


def normalize_alias(text: str) -> str:
    return " ".join((text or "").strip().split()).casefold()


def parse_best_of(value: Any) -> int:
    text = str(value or "").upper().replace("BO", "").strip()
    return int(text or 0)


def player_column_sort(column: str) -> tuple[str, int]:
    match = PLAYER_COLUMN_RE.match(column)
    if not match:
        return ("z", 0)
    return (match.group(1), int(match.group(2)))


def player_columns_from_row(row: dict[str, Any]) -> list[str]:
    columns = [column for column in row if PLAYER_COLUMN_RE.match(str(column))]
    return sorted(columns, key=player_column_sort) or DEFAULT_PLAYER_COLUMNS


def row_hash(row: dict[str, Any]) -> str:
    columns = ["match_date", "event_name", "stage", *player_columns_from_row(row), "score_a", "score_b", "best_of", "winning_side"]
    payload = {column: str(row.get(column, "") or "").strip() for column in columns}
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()


def get_or_create_player(canonical_name: str, source: str = "manual", alias_text: str | None = None) -> int:
    canonical_name = " ".join(canonical_name.strip().split())
    if not canonical_name:
        raise ValueError("Canonical player name is required")
    alias_text = alias_text or canonical_name

    with get_db() as db:
        existing = db.execute(
            "SELECT id FROM players WHERE canonical_name = ? COLLATE NOCASE",
            (canonical_name,),
        ).fetchone()
        if existing:
            player_id = int(existing["id"])
        else:
            player_id = int(
                db.execute(
                    "INSERT INTO players(canonical_name) VALUES (?)",
                    (canonical_name,),
                ).lastrowid
            )
        db.execute(
            """
            INSERT OR IGNORE INTO player_aliases(player_id, source_system, alias_text, approved)
            VALUES (?, ?, ?, 1)
            """,
            (player_id, source, alias_text.strip()),
        )
        return player_id


def resolve_player(alias_text: str, source: str = "owcs") -> int | None:
    alias_text = " ".join((alias_text or "").strip().split())
    if not alias_text:
        return None
    with get_db() as db:
        row = db.execute(
            """
            SELECT p.id
            FROM player_aliases pa
            JOIN players p ON p.id = pa.player_id
            WHERE pa.approved = 1 AND pa.alias_text = ? COLLATE NOCASE
            ORDER BY CASE WHEN pa.source_system = ? THEN 0 ELSE 1 END
            LIMIT 1
            """,
            (alias_text, source),
        ).fetchone()
        if row:
            return int(row["id"])
        player = db.execute(
            "SELECT id FROM players WHERE canonical_name = ? COLLATE NOCASE",
            (alias_text,),
        ).fetchone()
        return int(player["id"]) if player else None


def validate_match(payload: dict[str, Any], side_a: list[int], side_b: list[int]) -> list[str]:
    errors: list[str] = []
    if len(side_a) < 5 or len(side_b) < 5:
        errors.append("Each side must have at least five resolved players.")
    if len(set(side_a)) != len(side_a) or len(set(side_b)) != len(side_b):
        errors.append("A side contains duplicate players.")
    if set(side_a).intersection(side_b):
        errors.append("The same player appears on both sides.")
    if payload["side_a_score"] < 0 or payload["side_b_score"] < 0:
        errors.append("Scores cannot be negative.")
    if payload["best_of"] <= 0:
        errors.append("Best-of must be a positive number.")
    if payload["winning_side"] not in {"a", "b", "draw"}:
        errors.append("Winning side must be a, b, or draw.")
    if payload["winning_side"] == "a" and payload["side_a_score"] <= payload["side_b_score"]:
        errors.append("Side A winner must have a higher score.")
    if payload["winning_side"] == "b" and payload["side_b_score"] <= payload["side_a_score"]:
        errors.append("Side B winner must have a higher score.")
    return errors


def create_match(payload: dict[str, Any], side_a: list[int], side_b: list[int], rebuild: bool = True) -> int:
    errors = validate_match(payload, side_a, side_b)
    if errors:
        raise ValueError("; ".join(errors))

    with get_db() as db:
        if payload.get("source_match_id"):
            existing = db.execute(
                "SELECT id FROM matches WHERE source_system = ? AND source_match_id = ?",
                (payload["source_system"], payload["source_match_id"]),
            ).fetchone()
            if existing:
                raise ValueError("Duplicate source match ID.")

        match_id = int(
            db.execute(
                """
                INSERT INTO matches(
                    source_system, source_match_id, match_datetime, competition_name,
                    competition_type, best_of, side_a_score, side_b_score, winning_side,
                    status, raw_payload_json, row_hash
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    payload["source_system"],
                    payload.get("source_match_id"),
                    payload["match_datetime"],
                    payload["competition_name"],
                    payload.get("competition_type"),
                    payload["best_of"],
                    payload["side_a_score"],
                    payload["side_b_score"],
                    payload["winning_side"],
                    payload.get("status", "approved"),
                    payload.get("raw_payload_json"),
                    payload.get("row_hash"),
                ),
            ).lastrowid
        )
        for player_id in side_a:
            db.execute("INSERT INTO match_rosters(match_id, player_id, side) VALUES (?, ?, 'a')", (match_id, player_id))
        for player_id in side_b:
            db.execute("INSERT INTO match_rosters(match_id, player_id, side) VALUES (?, ?, 'b')", (match_id, player_id))

    if rebuild:
        rebuild_ratings("match_edit", f"Imported match {match_id}")
    return match_id


def approve_match(match_id: int) -> None:
    with get_db() as db:
        db.execute("UPDATE matches SET status = 'approved', updated_at = CURRENT_TIMESTAMP WHERE id = ?", (match_id,))
    rebuild_ratings("match_edit", f"Approved match {match_id}")


def reject_match(match_id: int) -> None:
    with get_db() as db:
        db.execute("UPDATE matches SET status = 'rejected', updated_at = CURRENT_TIMESTAMP WHERE id = ?", (match_id,))
    rebuild_ratings("match_edit", f"Rejected match {match_id}")


def sheet_rows(sheet: Worksheet) -> tuple[list[str], list[dict[str, Any]], dict[str, int]]:
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    index = {header: pos + 1 for pos, header in enumerate(headers) if header}
    rows = []
    for offset, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(values):
            continue
        rows.append({"_excel_row": offset, **{header: values[pos] if pos < len(values) else None for pos, header in enumerate(headers)}})
    return headers, rows, index


def sync_aliases_from_workbook(path: Path) -> int:
    workbook = load_workbook(path)
    if "Player Alias Map" not in workbook.sheetnames:
        return 0
    _, rows, _ = sheet_rows(workbook["Player Alias Map"])
    count = 0
    for row in rows:
        approved = str(row.get("approved") or "yes").strip().lower() in {"yes", "y", "true", "1"}
        if not approved:
            continue
        canonical = str(row.get("canonical_player") or "").strip()
        alias = str(row.get("alias_text") or "").strip()
        source = str(row.get("source") or "manual").strip() or "manual"
        if canonical and alias:
            get_or_create_player(canonical, source, alias)
            count += 1
    return count


def import_owcs_workbook() -> dict[str, int | str]:
    config = load_config()
    path = project_path(config["paths"]["workbook"])
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    alias_count = sync_aliases_from_workbook(path)
    workbook = load_workbook(path)
    if "Match Entry" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the Match Entry sheet.")

    sheet = workbook["Match Entry"]
    headers, rows, header_index = sheet_rows(sheet)
    if "import_error" not in header_index:
        sheet.cell(row=1, column=len(headers) + 1, value="import_error")
        header_index["import_error"] = len(headers) + 1
    if "row_hash" not in header_index:
        sheet.cell(row=1, column=len(headers) + 2, value="row_hash")
        header_index["row_hash"] = len(headers) + 2

    imported = 0
    rejected = 0
    seen = 0
    for row in rows:
        seen += 1
        excel_row = int(row["_excel_row"])
        status = str(row.get("status") or "").strip().lower()
        digest = row_hash(row)
        source_match_id = f"owcs:{digest[:16]}"
        errors: list[str] = []

        if status not in {"ready", "approved"}:
            sheet.cell(row=excel_row, column=header_index["import_error"], value="Skipped: status is not ready or approved")
            sheet.cell(row=excel_row, column=header_index["row_hash"], value=digest)
            continue

        side_a: list[int] = []
        side_b: list[int] = []
        for column in player_columns_from_row(row):
            raw_name = str(row.get(column) or "").strip()
            if not raw_name:
                continue
            player_id = resolve_player(raw_name, "owcs")
            if player_id is None:
                errors.append(f"Unresolved alias in {column}: {row.get(column)}")
            elif PLAYER_COLUMN_RE.match(column).group(1) == "a":
                side_a.append(player_id)
            else:
                side_b.append(player_id)

        try:
            match_datetime = row.get("match_date")
            if isinstance(match_datetime, datetime):
                match_datetime_text = match_datetime.isoformat()
            else:
                match_datetime_text = datetime.fromisoformat(str(match_datetime)).isoformat()
            payload = {
                "source_system": "owcs",
                "source_match_id": source_match_id,
                "match_datetime": match_datetime_text,
                "competition_name": str(row.get("event_name") or "OWCS"),
                "competition_type": str(row.get("stage") or ""),
                "best_of": parse_best_of(row.get("best_of")),
                "side_a_score": int(row.get("score_a") or 0),
                "side_b_score": int(row.get("score_b") or 0),
                "winning_side": str(row.get("winning_side") or "").strip().lower(),
                "status": "approved" if status == "approved" else "pending",
                "raw_payload_json": json.dumps(row, default=str),
                "row_hash": digest,
            }
            errors.extend(validate_match(payload, side_a, side_b))
        except Exception as exc:
            errors.append(str(exc))

        with get_db() as db:
            existing = db.execute(
                "SELECT id, row_hash FROM matches WHERE source_system = 'owcs' AND source_match_id = ?",
                (source_match_id,),
            ).fetchone()
        if existing:
            sheet.cell(row=excel_row, column=header_index["import_error"], value="")
            sheet.cell(row=excel_row, column=header_index["row_hash"], value=digest)
            continue

        if errors:
            rejected += 1
            message = "; ".join(errors)
            sheet.cell(row=excel_row, column=header_index["import_error"], value=message)
            sheet.cell(row=excel_row, column=header_index["row_hash"], value=digest)
            with get_db() as db:
                db.execute(
                    "INSERT INTO import_issues(source_system, reference, message) VALUES ('owcs', ?, ?)",
                    (f"row {excel_row}", message),
                )
            continue

        create_match(payload, side_a, side_b, rebuild=False)
        sheet.cell(row=excel_row, column=header_index["import_error"], value="")
        sheet.cell(row=excel_row, column=header_index["row_hash"], value=digest)
        imported += 1

    workbook.save(path)
    if imported:
        rebuild_ratings("match_edit", "OWCS workbook import")

    with get_db() as db:
        db.execute(
            """
            INSERT INTO sync_runs(source_system, status, items_seen, items_inserted, items_updated, finished_at)
            VALUES ('owcs', 'success', ?, ?, ?, CURRENT_TIMESTAMP)
            """,
            (seen, imported, alias_count),
        )
    return {"rows_seen": seen, "rows_imported": imported, "rows_rejected": rejected, "aliases_synced": alias_count}
